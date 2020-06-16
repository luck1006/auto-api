# -*- coding: UTF-8 -*-
import requests
import re
import openpyxl
from bs4 import BeautifulSoup
from app.actions.tools.queryFromSql import queryFromSql
from string import digits

# s = 'abc123def456ghi789zero0'
# remove_digits = str.maketrans('', '', digits)
# res = s.translate(remove_digits)
# print(res)

remove_digits = str.maketrans('', '', digits)


def oa_member():
    '''
    建立部门人员名称带不带数字的对应关系 希望一个部门不要有重名的哈哈😀
    :return:
    '''
    sql = "select name_cn from t_oa_user where dept_fir = '系统产品研发BU'"
    d = queryFromSql('jira', sql).querySql()
    member = {}
    for i in d[1:]:
        for y in i:
            res = y.translate(remove_digits)
            member[res] = y
    return member


def wiki_team():
    wiki = 'http://wiki.tuniu.org/pages/viewpage.action?pageId=111032223'
    try:
        res = requests.get(wiki)
    except Exception as e:
        return {}
    soup = BeautifulSoup(res.text, 'lxml')
    team_dic = {}
    member = oa_member()
    for idx, tr in enumerate(soup.find_all('tr')):
        tds = tr.find_all('td')
        temp = [i.text.replace('\xa0', '') for i in tds]
        if len(temp) == 10:
            aaa = {}
            aaa['po'] = temp[6]
            aaa['dev'] = temp[7]
            aaa['tester'] = temp[8]
            # 标准 名字间、分隔 如： 华金良、郭喜芝、李秋婷、赵君、朱家标、刘桃桃
            # 下边是因为文档不规范做了一些处理，如果后边出了问题 把文档里不规范的东西删掉

            # 干扰字段
            rep = ['前端', '服务端', ':', '：', '，', ' ', '后台研发', '后端', '中台敏捷', '订单中台']
            member_list = []
            for x in aaa:
                for i in rep:
                    aaa[x] = aaa[x].replace(i, '、')
                nums = aaa[x].split('、')
                for y in nums:
                    one = re.split('（|\(', y)[0]
                    # 万一wiki上边带了数字,先去掉
                    one = one.translate(remove_digits)
                    try:
                        member_list.append(member[one])
                    except Exception as e:
                        print("未在名单中：" + str(e))
            team_dic[temp[2]] = member_list
    # print(team_dic)
    return team_dic


if __name__ == '__main__':


    '同步名单中迭代组成员'
    excel_path = '../../../views/TNJ/名单.xlsx'
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb["Sheet1"]
    wb.remove(sheet)
    wb.save(excel_path)
    wb.create_sheet("Sheet1")
    sheet = wb["Sheet1"]
    sheet.cell(row=1, column=1).value = '姓名'
    sheet.cell(row=1, column=2).value = 'ScrumTeam'

    max_column = sheet.max_column
    max_row = sheet.max_row
    mem = wiki_team()
    data = []
    for i, v in mem.items():
        for y in v:
            sheet.cell(row=max_row + 1, column=1).value = y
            sheet.cell(row=max_row + 1, column=2).value = i
            max_row += 1
    wb.save(excel_path)
