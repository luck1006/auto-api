# -*- coding: utf-8 -*-
# TIME:         7.25
# Author:       huangfang
# Explain：     获取上周五到本周四的资本化信息并入库
import os,sys

sys.path.append('../../')
os.path.abspath('../../../')
from app.actions.tools import timer
import openpyxl
from openpyxl import Workbook

import xlrd
import xlwt
from xlutils.copy import copy
from app.actions.tools.queryFromSql import queryFromSql
import pandas as pd
from  sqlalchemy import create_engine,func
import pymysql
from app.actions.tools.conf import dbconf

from app.actions.tools.SaveInit import SaveInit
from app.actions.tools.DbUtility import *
from app import db

import datetime
from app.models.models import *
from config import *
from app.actions.tools import ddingForJira
from app.actions.tools import fb







def jira_tuniu_test(start_time, end_time):
    #自定义时间：
    #start_time = '2019-08-09'
    #end_time = '2019-08-15'
    #start_time, end_time, start, end, day = timer.time_l()
    year = int(end_time.split('-')[0])
    month = int(end_time.split('-')[1])
    day = int(end_time.split('-')[2])
    week = str(datetime.date(year, month, day).isocalendar()[1]) + 'W'
    db = 'jira'
    excelWriter = pd.ExcelWriter('OnlineBU'+'-'+week+'.xlsx', engine='openpyxl')

    # 只是为了部门在第一个sheet
    l = pd.DataFrame(['', ''])
    l.to_excel(excelWriter, sheet_name='部门人员统计')

    # 查询worklog表中数据
    sql = f'''
    SELECT  b.name_cn AS 姓名, b.dept_sec AS 组织架构, STR_TO_DATE(STARTDATE, '%Y-%m-%d') AS 日期
    	,a.timeworked / 3600 / 8 AS 工时
    	, CONCAT(d.pkey, '-', c.issuenum) AS 单号
    	, e.pname AS 类型, c.id, g.LINKNAME
    	, IF(e.pname IN ('技术需求', '业务需求', '新技术支持'), c.id, f.SOURCE) AS SOURCE
    FROM worklog a
    	LEFT JOIN t_oa_user b ON a.AUTHOR = b.name_en
    	LEFT JOIN jiraissue c ON c.id = a.issueid
    	LEFT JOIN project d ON c.PROJECT = d.ID
    	LEFT JOIN issuetype e ON c.issuetype = e.id
    	LEFT JOIN issuelink f ON f.DESTINATION = c.id
    	LEFT JOIN issuelinktype g ON f.LINKTYPE = g.id
    WHERE b.dept_fir = '系统产品研发BU'
    	AND b.dept_sec in ("研发TeamI","研发TeamJ","研发TeamK","研发TeamL","研发TeamM","设计TeamN","产品TeamP","质量TeamS")
    	AND e.pname IN ('子事务', 'Sub-task')
    	AND g.id = 10100
     -- 取jira为('子事务', 'Sub-task') 链接类型是jira_subtask_link
    	AND STARTDATE BETWEEN '{start_time}' AND '{end_time}'
    UNION ALL
    SELECT b.name_cn AS 姓名, b.dept_sec AS 组织架构, STR_TO_DATE(STARTDATE, '%Y-%m-%d') AS 日期
    	, a.timeworked / 3600 / 8 AS 工时
    	, CONCAT(d.pkey, '-', c.issuenum) AS 单号
    	, e.pname AS 类型, c.id, g.LINKNAME
    	, IF(e.pname IN ('技术需求', '业务需求', '新技术支持'), c.id, f.SOURCE) AS SOURCE
    	-- 限制父级jira在('技术需求', '业务需求', '新技术支持')
    FROM worklog a
    	LEFT JOIN t_oa_user b ON a.AUTHOR = b.name_en
    	LEFT JOIN jiraissue c ON c.id = a.issueid
    	LEFT JOIN project d ON c.PROJECT = d.ID
    	LEFT JOIN issuetype e ON c.issuetype = e.id
    	LEFT JOIN issuelink f ON f.DESTINATION = c.id
    	LEFT JOIN issuelinktype g ON f.LINKTYPE = g.id
    WHERE b.dept_fir = '系统产品研发BU'
    	AND b.dept_sec in ("研发TeamI","研发TeamJ","研发TeamK","研发TeamL","研发TeamM","设计TeamN","产品TeamP","质量TeamS")
    	AND e.pname NOT IN ('子事务', 'Sub-task')
        AND STARTDATE BETWEEN '{start_time}' AND '{end_time}'
	-- 取其他jira 此时不限制
    '''

    data = queryFromSql(db, sql).querySql()
    res = pd.DataFrame(data[1:], columns=data[0])
    ids = res['SOURCE']
    ids = ids.tolist()
    l = list(set(ids))
    ids = ''
    for i in l:
        if i != "":
            ids += i + ','
    ids = ids.strip().strip(',')

    # 查询jra父级相关属性或本身 ('业务需求','技术需求','新技术支持')属性
    sql_2 = f'''
    SELECT distinct c.id AS SOURCE,  CONCAT(d.pkey, '-', c.issuenum) AS 父级jira
        , e.pname AS 类型
    --     , f.SOURCE 父级id
        , c.SUMMARY 主题
        , c.DESCRIPTION  描述,
        (SELECT b.name_cn FROM customfieldvalue af left join t_oa_user b on b.name_en = af.STRINGVALUE WHERE af.ISSUE=c.ID AND CUSTOMFIELD=10023) AS 需求提出人,
        (SELECT cfd.customvalue FROM customfieldvalue af LEFT JOIN customfieldoption cfd ON cfd.id=af.STRINGVALUE WHERE af.ISSUE=c.ID AND af.CUSTOMFIELD=17905) AS 需求提出部门,
        IFNULL((SELECT cfop.customvalue FROM customfieldvalue cfvp,customfieldoption cfop WHERE c.ID= cfvp.ISSUE AND cfvp.CUSTOMFIELD = 12100 AND cfop.ID = cfvp.STRINGVALUE),'未立项()') AS 立项项目名称,
        (SELECT STRINGVALUE FROM customfieldvalue af WHERE af.ISSUE = c.ID AND CUSTOMFIELD = 10304) Epic
    FROM jiraissue c
        LEFT JOIN  worklog a ON c.id = a.issueid
        LEFT JOIN t_oa_user b ON a.AUTHOR = b.name_en
        LEFT JOIN project d ON c.PROJECT = d.ID
        LEFT JOIN issuetype e ON c.issuetype = e.id
        LEFT JOIN issuelink f ON f.DESTINATION = c.id
    WHERE c.id in ({ids}) AND e.pname in ('业务需求','技术需求','新技术支持')
    '''

    # 序号	ID	LINKNAME	INWARD	OUTWARD	pstyle
    # 1	10000	Blocks	is blocked by	blocks
    # 2	10001	Cloners	is cloned by	clones
    # 3	10002	Duplicate	is duplicated by	duplicates
    # 4	10003	Relates	relates to	relates to
    # 5	10100	jira_subtask_link	jira_subtask_inward	jira_subtask_outward	jira_subtask
    # 6	10200	Epic-Story Link	has Epic	is Epic of	jira_gh_epic_story

    data_2 = queryFromSql(db, sql_2).querySql()
    res_2 = pd.DataFrame(data_2[1:], columns=data_2[0])
    df = pd.DataFrame.merge(res, res_2, how='left', on='SOURCE')

    # 取所有的填写工时的人
    members = tuple(set(df['姓名']))

    # 工时列 字符串转数字
    df[['工时']] = df[['工时']].apply(pd.to_numeric)

    df.to_excel(excelWriter, sheet_name='明细', index=None)
    excelWriter.save()
    #print('df11111====',type(df),df)



    # 从excel中取数据
    wb2 = openpyxl.load_workbook('OnlineBU'+'-'+week+'.xlsx')
    sheet = wb2["明细"]
    # excel 行数 列数
    max_column = sheet.max_column
    max_row = sheet.max_row

    # 查询部门当中时间段内未填写工时的人员
    sql_mem_0 = f'''
    select name_cn as 姓名,dept_sec as 组织架构  from t_oa_user where dept_fir ='系统产品研发BU'  AND dept_sec in ("研发TeamI","研发TeamJ","研发TeamK","研发TeamL","研发TeamM","设计TeamN","产品TeamP","质量TeamS")  and level_cn  !='已离职'
     and name_cn not in {members}
     order by dept_sec, name_cn
    '''

    data = queryFromSql(db, sql_mem_0).querySql()
    #print('data1111====',type(data),data)

    # 未填写工时的人加入excel
    for i in data[1:]:
        for y in range(0, len(i)):
            sheet.cell(max_row + 1, y + 1).value = i[y]
            # 默认立项项目名称 未立项()
            sheet.cell(max_row + 1, 16).value = '未立项()'
            # 默认工时0
            sheet.cell(max_row + 1, 4).value = '0'
        max_row += 1
    wb2.save('OnlineBU'+'-'+week+'.xlsx')

    df = pd.read_excel(excelWriter, sheet_name='明细')
    df = df.sort_values(by=['姓名', '日期'])

    #print('df22222====',type(df),df)
    #print("value=2222===,",df.get_values())
    df1 = df.groupby(by=['立项项目名称', '姓名'])['工时'].sum()
    df1 = df1.groupby(level=0, group_keys=False).nlargest(200)
    df1.to_excel(excelWriter, sheet_name='项目预估与投入填写')
    #df.to_excel(excelWriter, sheet_name='明细', index=None)

    df2 = df.groupby(by=['组织架构', '姓名'])['工时'].sum()
    df2 = df2.groupby(level=0, group_keys=False).nlargest(200)
    df2.to_excel(excelWriter, sheet_name='部门人员统计')



    excelWriter.save()
    excelWriter.close()
    return df





if __name__ == '__main__':
    #每次清除detail表数据
    save = SaveInit()
    save.connect()
    detail_sql = "truncate table time_working_detail"
    save.insert(detail_sql)
    save.close()
    print('清除time_working_detail表数据成功')

    start_time, end_time, start, end, day = timer.time_l()
    #转成str类型
    start_time=start_time.strftime('%Y-%m-%d')
    end_time=end_time.strftime('%Y-%m-%d')
    #自定义开始结束时间
    # start_time = '2020-1-10'
    # end_time = '2020-1-16'
    dftest = jira_tuniu_test(start_time, end_time)
    pymysql.install_as_MySQLdb()
    #数据库配置
    dbHost=dbconf["DB_HOST"]
    dbUser = dbconf["DB_USER"]
    dbPass = dbconf["DB_PASS"]
    dbName = dbconf["DB_NAME"]
    dbPort = dbconf["DB_PORT"]
    sqllink= 'mysql+mysqldb://'+dbUser+':'+dbPass+'@'+dbHost+':'+dbPort+'/'+dbName+'?charset=utf8'


    #conn=create_engine('mysql+mysqldb://root:@MonLey880124@10.28.32.130:3306/pandora?charset=utf8')

    conn=create_engine(sqllink)
    dftest.to_sql('time_working_detail', conn, schema='pandora', if_exists='append', index=True)
    print ('资本化明细数据入库成功')
    #获取当前统计数据的周数
    year = int(end_time.split('-')[0])
    month = int(end_time.split('-')[1])
    day = int(end_time.split('-')[2])
    week = str(datetime.date(year, month, day).isocalendar()[1]) + 'W'
    print ('资本化数据统计周期为',week)
    print('资本化数据入库按周统计的表：time_working')
    save = SaveInit()
    save.connect()
    #计算当周的项目工时
    # 项目总工时
    projecttime = db.session.query(func.sum(Time_working_detail.工时)).all()
    # 项目中部门"资本化"工时
    # projecttime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(
    #     Time_working_detail.立项项目名称.in_(['常规度假转化-Q4','老客运营，流量获取-Q4','聚合频道转化-Q4','新客转化-Q4','供应链系统维护项目-Q4','零售订单转化-Q4','框架工具-Q4','打包预定转化-Q4','大数据-Q4','会员，营销，订单系统维护项目-Q4'])).all()
    projecttime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(Time_working_detail.立项项目名称 !='未立项()').all()
    # 项目中【未立项】的工时
    projecttime2 = db.session.query(func.sum(Time_working_detail.工时)).filter(
        Time_working_detail.立项项目名称 == '未立项()').all()
    # 项目工时
    print(projecttime, projecttime1, projecttime2)
    print(projecttime[0][0],projecttime1[0][0],projecttime2[0][0])

    # if projecttime is not None:
    #     totaltime = projecttime[0][0]
    # else:
    #     totaltime = 0
    if projecttime1 is not None:
        projectworktime = projecttime1[0][0]
    else:
        projectworktime = 0
    if projecttime2 is not None:
        unprojectworktime = projecttime2[0][0]
    else:
        unprojectworktime = 0

    # projectTime1 = projecttime[0][0] - projecttime1[0][0] - projecttime2[0][0]
    projectTime1 = projecttime[0][0] - projectworktime - unprojectworktime  #  dff2020-02-11 1M
    projectTime ='%.2f' %projectTime1
    print("projectTime==",projectTime)
    print(type(projectTime))
    #计算当周的技术支持工时
    supportTime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(Time_working_detail.类型_y == '新技术支持').all()
    supportTime ='%.2f' %(supportTime1[0][0])
    #计算当周的理论工时（总人数*5）
    totalnum = db.session.query(Time_working_detail.姓名).distinct().count()
    theoreticalTime = '%.2f' %(totalnum * 5)
    # 计算当周的资本化工时
    capitalizationTime1 =projecttime[0][0] - supportTime1[0][0] - projectTime1
    capitalizationTime ='%.2f' %capitalizationTime1
    #未立项工时平摊到各个模块
    # weilixiangTime=projecttime2[0][0]/10
    weilixiangTime = unprojectworktime / 10    #  dff2020-02-11 1M
    #print ('weilixiangTime====',weilixiangTime)
    #计算各个资本化模块的工时(10个)

    # 【资本化】零售
    lingshouTime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(
        Time_working_detail.立项项目名称.like('零售%')).all()
    lingshouTimet=lingshouTime1[0][0]
    if lingshouTimet is not None:
       lingshouTime = '%.2f' %(lingshouTime1[0][0]+weilixiangTime)
    else:
        lingshouTime = '%.2f' % (0 + weilixiangTime)

    # 【资本化】度假转化
    zhuanhuaTime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(
        Time_working_detail.立项项目名称.like('常规度假转化%')).all()
    zhuanhuaTimet= zhuanhuaTime1[0][0]
    if zhuanhuaTimet is not None:
       zhuanhuaTime = '%.2f' %(zhuanhuaTime1[0][0]+weilixiangTime)
       print("zhuanhuaTime==", zhuanhuaTime)

    else:
        zhuanhuaTime = '%.2f' % (0 + weilixiangTime)

    # 【资本化】新客转化
    xinkeTime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(
        Time_working_detail.立项项目名称.like('新客转化%')).all()
    xinkeTimet=xinkeTime1[0][0]
    if xinkeTimet is not None:
        xinkeTime = '%.2f' %(xinkeTime1[0][0]+weilixiangTime)
    else:
        xinkeTime = '%.2f' %(0+weilixiangTime)

    # 【资本化】老客运营
    yunyingTime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(
        Time_working_detail.立项项目名称.like('老客运营%')).all()
    yunyingTimet=yunyingTime1[0][0]
    if yunyingTimet is not None:
        yunyingTime = '%.2f' %(yunyingTime1[0][0]+weilixiangTime)
    else:
        yunyingTime = '%.2f' % (0 + weilixiangTime)

    # 【资本化】架构
    jiagouTime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(
        Time_working_detail.立项项目名称.like('框架%')).all()
    jiagouTimet=jiagouTime1[0][0]
    if jiagouTimet is not None:
        jiagouTime = '%.2f' %(jiagouTime1[0][0]+weilixiangTime)
    else:
        jiagouTime = '%.2f' % (0 + weilixiangTime)

     # 【资本化】打包预定
    dabaoTime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(
            Time_working_detail.立项项目名称.like('打包预定%')).all()
    dabaoTimet = dabaoTime1[0][0]
    if dabaoTimet is not None:
            dabaoTime = '%.2f' % (dabaoTime1[0][0] + weilixiangTime)
    else:
        dabaoTime = '%.2f' % (0 + weilixiangTime)
    print('dabaoTime====',dabaoTime)


            # 【资本化】聚合
    juheTime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(
            Time_working_detail.立项项目名称.like('聚合%')).all()
    juheTimet = dabaoTime1[0][0]
    if juheTimet is not None:
        juheTime = '%.2f' % (juheTime1[0][0] + weilixiangTime)
    else:
        juheTime = '%.2f' % (0 + weilixiangTime)
    print('juheTime====',juheTime)

    # 【资本化】会员，订单
    huiyuanTime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(
        Time_working_detail.立项项目名称.like('会员%')).all()
    huiyuanTimet=huiyuanTime1[0][0]
    if huiyuanTimet is not None:
       huiyuanTime = '%.2f' %(huiyuanTime1[0][0]+weilixiangTime)
    else:
        huiyuanTime = '%.2f' % (0 + weilixiangTime)

    # 【资本化】供应链
    gongyinglianTime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(
        Time_working_detail.立项项目名称.like('供应链%')).all()
    gongyinglianTimet=gongyinglianTime1[0][0]
    if gongyinglianTimet is not None:
       gongyinglianTime = '%.2f' %(gongyinglianTime1[0][0]+weilixiangTime)
    else:
        gongyinglianTime = '%.2f' % (0 + weilixiangTime)

    # 【资本化】大数据
    dashujuTime1 = db.session.query(func.sum(Time_working_detail.工时)).filter(
        Time_working_detail.立项项目名称.like('大数据%')).all()
    dashujuTimet=dashujuTime1[0][0]
    if dashujuTimet is not None:
       dashujuTime = '%.2f' %(dashujuTime1[0][0]+weilixiangTime)
    else:
        dashujuTime = '%.2f' % (0 + weilixiangTime)

    del_sql = "delete from time_working WHERE WEEK='%s'" % week

    save.insert(del_sql)
    sql = "INSERT INTO time_working VALUES('%s',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)" \
          %(week,capitalizationTime,projectTime,supportTime,theoreticalTime,totalnum,lingshouTime,zhuanhuaTime,xinkeTime,\
            yunyingTime,jiagouTime,dabaoTime,juheTime,huiyuanTime,gongyinglianTime,dashujuTime,str(datetime.datetime.now().year)+str(datetime.datetime.now().month))
    print (sql)
    save.insert(sql)

    #print (sql)
    save.close()
    print('资本化数据入周统计表time_working成功')
    #在excel表中新增加几个sheet
    #打开excel表
    wb3 = openpyxl.load_workbook('OnlineBU'+'-'+week+'.xlsx')
    #删除明细sheet
    ws = wb3["明细"]
    wb3.remove(ws)
    #增加sheet
    Sheet1 = wb3.create_sheet('概述(填写)',0)
    Sheet2 = wb3.create_sheet('在行项目进展(填写)',2)
    Sheet3 = wb3.create_sheet('需求&筹备项目(填写)',3)
    #Sheet4 = wb3.create_sheet('填写说明',5)
    wb3.save('OnlineBU'+'-'+week+'.xlsx')
   ##在sheet1中添加数据
    wb3 = openpyxl.load_workbook('OnlineBU'+'-'+week+'.xlsx')




    #打印所有的sheet名称
    #print(wb3.sheetnames)
    #wb3 = Workbook()
    wb4 = wb3.active  # 激活sheet
    # 表头
    wb4.cell(1, 1, '线上BU')
    #往sheet中的第二行第二列写入需要统计项的名称
    list=['周期(周)','部门总人数','项目统计人数','本期参与项目人数','部门理论人天','筹备期项目实际人天','在行项目实际人天','项目理论人天','项目人员利用率','项目人天占比','小需求每周实际人天','小需求人天占比','技术支持每周实际人天','技术支持人天占比']
    #填入相应的数据
    totalnump=totalnum
    choubeitime=0
    liyonglip=round(((choubeitime+float(projectTime))/(totalnum*5))*100)
    liyonglip1=str(liyonglip)+'%'
    zhanbip=round(((choubeitime+float(projectTime))/(totalnump*5))*100)
    zhanbip1=str(zhanbip)+'%'
    zhanbixuqiu=round(((float(capitalizationTime))/(totalnum*5))*100)
    zhanbixuqiu1=str(zhanbixuqiu)+'%'
    zhanbij=round(((float(supportTime))/(totalnump*5))*100)
    zhanbij1=str(zhanbij)+'%'
    #print("zhanbij==",liyonglip,zhanbip,zhanbixuqiu,zhanbij)
    list2=[week,totalnum,totalnump,totalnump,totalnum*5,choubeitime,float(projectTime),totalnump*5,liyonglip1,zhanbip1,float(capitalizationTime),zhanbixuqiu1,float(supportTime),zhanbij1]
    for i in range(len(list)):
     wb4.cell(2, i+1, list[i])
    for j in range(len(list2)):
     wb4.cell(3, j + 1, list2[j])
    wb3.save('OnlineBU' + '-' + week + '.xlsx')  # 保存
    #打开sheet1
    sheet = wb3.get_sheet_by_name('概述(填写)')
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 15
    sheet.column_dimensions['I'].width = 15
    sheet.column_dimensions['J'].width = 15
    sheet.column_dimensions['K'].width = 20
    sheet.column_dimensions['L'].width = 15
    sheet.column_dimensions['M'].width = 20
    sheet.column_dimensions['N'].width = 20


    #打开sheet3
    sheet = wb3.get_sheet_by_name('在行项目进展(填写)')
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 15
    sheet.column_dimensions['I'].width = 15
    sheet.column_dimensions['J'].width = 15
    sheet.column_dimensions['K'].width = 15
    sheet.column_dimensions['L'].width = 15
    sheet.column_dimensions['M'].width = 15
    sheet.column_dimensions['O'].width = 20
    sheet.column_dimensions['Q'].width = 25
    sheet.column_dimensions['R'].width = 15
    sheet.column_dimensions['S'].width = 20
    sheet.column_dimensions['T'].width = 20

    #填写在行项目情况
    sheet['A1'] = '在行项目基本信息'
    sheet['A2'] = '项目状态'
    sheet['B2'] = '项目编号'
    sheet['C2'] = '项目名称'
    sheet['D2'] = '项目经理'
    sheet['E2'] = '所属研发部门'
    sheet['F2'] = '项目星级'
    sheet['G2'] = '项目简介'
    sheet['H2'] = '需求提出部门'
    sheet['I2'] = '费用承担部门'
    sheet['J2'] = '立项月份'
    sheet['K2'] = '计划开始时间'
    sheet['L2'] = '计划上线时间'
    sheet['M2'] = '计划结项时间'
    sheet['N2'] = '时间进度'
    sheet['O2'] = '进度与计划匹配度'
    sheet['P2'] = '本周进度'
    sheet['Q2'] = '目前存在主要风险和应对措施'
    sheet['R2'] = '计划总人天'
    sheet['S2'] = '实际已投入人天'
    sheet['T2'] = '资源使用占比'

    #list3=['A','B','C','D']
    #for k in range(len(list3)):
        #print(list3[k])
        #S=('{1}{0}{1}'.format(list3[k], "'")).upper()
        #print ((S))
        #sheet.column_dimensions[S].width = 30

    # 打开sheet4
    sheet = wb3.get_sheet_by_name('需求&筹备项目(填写)')
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['J'].width = 15
    sheet.column_dimensions['K'].width = 15
    sheet.column_dimensions['L'].width = 15
    sheet.column_dimensions['M'].width = 15
    sheet.column_dimensions['N'].width = 15
    sheet.column_dimensions['P'].width = 15

    # 填写筹备项目情况
    sheet['A1'] = '业务需求：公司业务部门转来的需求'
    sheet['A2'] = '序号'
    sheet['B2'] = '周度'
    sheet['C2'] = '提出时间'
    sheet['D2'] = '提出部门'
    sheet['E2'] = '业务部门'
    sheet['F2'] = '需求/项目名称'
    sheet['G2'] = '需求/项目介绍'
    sheet['H2'] = '现状'
    sheet['I2'] = '预期收益'
    sheet['J2'] = '期望完成时间'
    sheet['K2'] = '计划总人天'
    sheet['L2'] = 'PMO会议结论'
    sheet['M2'] = '所属研发部门'
    sheet['N2'] = '研发责任人'
    sheet['O2'] = '状态'
    sheet['P2'] = '目前进展及风险'
    sheet['A3'] = 1
    sheet['B3'] = week
    sheet['C3'] = '2019-08-26'
    sheet['D3'] = '线上BU'
    sheet['E3'] = '线上BU'
    #打开sheet5
    sheet = wb3.get_sheet_by_name('项目预估与投入填写')
    sheet.column_dimensions['A'].width = 30

    wb3.save('OnlineBU'+'-'+week+'.xlsx')  # 保存
    #钉钉群发送
    url = fb.excel_url('OnlineBU'+'-'+week+'.xlsx')

    ddingForJira.send_dingding(url)




















